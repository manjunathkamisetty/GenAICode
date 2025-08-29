#!/usr/bin/env python3
"""
COBOL File Analyzer
===================
This script analyzes COBOL projects to extract statistics about:
- Number of COBOL programs
- Number of JCL (Job Control Language) files  
- Number of procedures/copybooks
- Number of input and output files referenced

Author: Generated for COBOL project analysis
"""

import os
import re
import json
import csv
from pathlib import Path
from collections import defaultdict, Counter
from datetime import datetime

# Optional dependencies for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class CobolFileAnalyzer:
    def __init__(self, root_directory=None):
        """Initialize the analyzer with a root directory to scan."""
        self.root_directory = Path(root_directory) if root_directory else Path('.')
        
        # File extension patterns for different file types
        self.file_patterns = {
            'cobol_programs': {
                'extensions': ['.cbl', '.cob', '.cobol', '.pgm'],
                'description': 'COBOL Programs'
            },
            'jcl_files': {
                'extensions': ['.jcl', '.job'],
                'description': 'JCL (Job Control Language) Files'
            },
            'copybooks': {
                'extensions': ['.cpy', '.copy', '.inc'],
                'description': 'Copybooks'
            },
            'procedures': {
                'extensions': ['.prc', '.proc', '.prn'],
                'description': 'Procedures'
            },
            'control_cards': {
                'extensions': ['.ctc', '.ctl', '.card', '.cc', '.ctrl'],
                'description': 'Control Cards (CTC)'
            },
            'data_files': {
                'extensions': ['.dat', '.data', '.txt', '.csv', '.seq'],
                'description': 'Data Files'
            }
        }
        
        # Patterns to identify input/output operations in COBOL code
        self.io_patterns = {
            'input_operations': [
                r'READ\s+(\w+)',
                r'ACCEPT\s+.*FROM\s+(\w+)',
                r'OPEN\s+INPUT\s+(\w+)',
                r'SELECT\s+(\w+)\s+ASSIGN\s+TO\s+.*INPUT'
            ],
            'output_operations': [
                r'WRITE\s+(\w+)',
                r'DISPLAY\s+.*UPON\s+(\w+)',
                r'OPEN\s+OUTPUT\s+(\w+)',
                r'SELECT\s+(\w+)\s+ASSIGN\s+TO\s+.*OUTPUT'
            ]
        }
        
        self.results = {
            'scan_timestamp': datetime.now().isoformat(),
            'root_directory': str(self.root_directory.absolute()),
            'file_counts': defaultdict(int),
            'folder_analysis': defaultdict(lambda: defaultdict(int)),
            'io_analysis': {
                'input_files': set(),
                'output_files': set(),
                'file_references': defaultdict(list)
            },
            'detailed_files': defaultdict(list),
            'jcl_datasets': [],  # New: Store JCL dataset information
            'excluded_counts': defaultdict(int)  # Track excluded files
        }

    def scan_directory(self):
        """Recursively scan the directory structure and analyze files."""
        print(f"Starting analysis of directory: {self.root_directory}")
        
        if not self.root_directory.exists():
            print(f"Directory {self.root_directory} does not exist!")
            return
            
        for root, dirs, files in os.walk(self.root_directory):
            root_path = Path(root)
            relative_path = root_path.relative_to(self.root_directory)
            
            print(f"Scanning folder: {relative_path}")
            
            for file in files:
                file_path = root_path / file
                self._analyze_file(file_path, relative_path)
        
        self._finalize_analysis()

    def _should_exclude_file(self, file_path):
        """Check if file should be excluded from analysis based on content."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read().upper()
                
                # Check for exclusion patterns
                if 'THIS JOB IS ALREADY ON LINUX' in content:
                    return True
                    
                return False
                
        except Exception as e:
            print(f"Error checking exclusion for file {file_path}: {e}")
            return False

    def _analyze_file(self, file_path, folder_path):
        """Analyze a single file and categorize it."""
        file_extension = file_path.suffix.lower()
        file_name = file_path.name
        
        # Categorize file by extension
        for category, config in self.file_patterns.items():
            if file_extension in config['extensions']:
                # Check if file should be excluded
                if self._should_exclude_file(file_path):
                    # Track excluded files but don't count them
                    self.results['excluded_counts'][category] += 1
                    print(f"   Excluding {file_name} - contains 'THIS JOB IS ALREADY ON LINUX'")
                    return
                
                # Count lines of code
                line_count = self._count_lines_of_code(file_path)
                
                self.results['file_counts'][category] += 1
                self.results['folder_analysis'][str(folder_path)][category] += 1
                
                file_info = {
                    'name': file_name,
                    'path': str(file_path.relative_to(self.root_directory)),
                    'size': file_path.stat().st_size if file_path.exists() else 0,
                    'lines_of_code': line_count['total_lines'],
                    'non_empty_lines': line_count['non_empty_lines'],
                    'comment_lines': line_count['comment_lines'],
                    'code_lines': line_count['code_lines']
                }
                
                self.results['detailed_files'][category].append(file_info)
                
                # If it's a COBOL program, analyze I/O operations
                if category == 'cobol_programs':
                    self._analyze_cobol_io(file_path)
                # If it's a JCL file or procedure, extract dataset information
                elif category in ['jcl_files', 'procedures']:
                    self._analyze_jcl_datasets(file_path)
                
                break
        else:
            # File doesn't match known patterns
            self.results['file_counts']['other_files'] += 1
            self.results['folder_analysis'][str(folder_path)]['other_files'] += 1

    def _count_lines_of_code(self, file_path):
        """Count lines of code in a file, distinguishing between comments, empty lines, and code."""
        line_stats = {
            'total_lines': 0,
            'non_empty_lines': 0,
            'comment_lines': 0,
            'code_lines': 0
        }
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    line_stats['total_lines'] += 1
                    stripped_line = line.strip()
                    
                    if not stripped_line:
                        # Empty line
                        continue
                    
                    line_stats['non_empty_lines'] += 1
                    
                    # Check for comment lines based on file type
                    file_extension = file_path.suffix.lower()
                    is_comment = False
                    
                    if file_extension in ['.cbl', '.cob', '.cobol', '.pgm', '.cpy', '.copy', '.inc', '.prc']:
                        # COBOL comments: * in column 7 or entire line starting with *
                        if (len(stripped_line) > 0 and stripped_line[0] == '*') or \
                           (len(line) > 6 and line[6] == '*'):
                            is_comment = True
                        # COBOL also uses // for comments in some dialects
                        elif stripped_line.startswith('//'):
                            is_comment = True
                    elif file_extension in ['.jcl', '.job', '.proc']:
                        # JCL comments: //* or lines starting with //
                        if stripped_line.startswith('//*'):
                            is_comment = True
                    elif file_extension in ['.ctc', '.ctl', '.card', '.cc', '.ctrl']:
                        # Control card comments: * at beginning, # for some systems, or //
                        if stripped_line.startswith('*') or stripped_line.startswith('#') or \
                           stripped_line.startswith('//') or stripped_line.startswith('//*'):
                            is_comment = True
                    
                    if is_comment:
                        line_stats['comment_lines'] += 1
                    else:
                        line_stats['code_lines'] += 1
                        
        except Exception as e:
            print(f"Error counting lines in file {file_path}: {e}")
        
        return line_stats

    def _analyze_cobol_io(self, file_path):
        """Analyze COBOL file for input/output operations."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read().upper()  # COBOL is case-insensitive
                
                # Find input operations
                for pattern in self.io_patterns['input_operations']:
                    matches = re.findall(pattern, content, re.IGNORECASE)
                    for match in matches:
                        self.results['io_analysis']['input_files'].add(match)
                        self.results['io_analysis']['file_references'][match].append({
                            'file': str(file_path.relative_to(self.root_directory)),
                            'operation': 'INPUT'
                        })
                
                # Find output operations
                for pattern in self.io_patterns['output_operations']:
                    matches = re.findall(pattern, content, re.IGNORECASE)
                    for match in matches:
                        self.results['io_analysis']['output_files'].add(match)
                        self.results['io_analysis']['file_references'][match].append({
                            'file': str(file_path.relative_to(self.root_directory)),
                            'operation': 'OUTPUT'
                        })
                        
        except Exception as e:
            print(f"Error analyzing COBOL file {file_path}: {e}")

    def _analyze_jcl_datasets(self, file_path):
        """Analyze JCL/PROC file to extract dataset names and DISP parameters."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
                
                current_job = None
                current_step = None
                current_proc = None
                line_number = 0
                current_dd_statement = ""
                current_dd_line_start = 0
                
                for line in lines:
                    line_number += 1
                    stripped_line = line.strip()
                    
                    # Skip empty lines and comments
                    if not stripped_line or stripped_line.startswith('//*'):
                        continue
                    
                    # Extract job name
                    if stripped_line.startswith('//') and ' JOB ' in stripped_line:
                        current_job = stripped_line.split()[0][2:]  # Remove '//' prefix
                        continue
                    
                    # Extract step name
                    if stripped_line.startswith('//') and ' EXEC ' in stripped_line:
                        current_step = stripped_line.split()[0][2:]  # Remove '//' prefix
                        # Check if it's calling a PROC
                        if 'PROC=' in stripped_line.upper():
                            proc_match = re.search(r'PROC=([^,\s]+)', stripped_line.upper())
                            if proc_match:
                                current_proc = proc_match.group(1)
                        continue
                    
                    # Extract PROC name
                    if stripped_line.startswith('//') and ' PROC ' in stripped_line:
                        current_proc = stripped_line.split()[0][2:]  # Remove '//' prefix
                        continue
                    
                    # Handle DD statements (including multi-line)
                    if stripped_line.startswith('//') and ' DD ' in stripped_line:
                        # Process any previous DD statement
                        if current_dd_statement:
                            self._process_complete_dd_statement(
                                current_dd_statement, file_path, current_dd_line_start, 
                                current_job, current_step, current_proc
                            )
                        
                        # Start new DD statement
                        current_dd_statement = stripped_line
                        current_dd_line_start = line_number
                        
                        # Check if this line ends with continuation
                        if not stripped_line.rstrip().endswith(','):
                            # Complete DD statement on single line
                            self._process_complete_dd_statement(
                                current_dd_statement, file_path, current_dd_line_start,
                                current_job, current_step, current_proc
                            )
                            current_dd_statement = ""
                    
                    # Handle continuation lines
                    elif (current_dd_statement and 
                          stripped_line.startswith('//') and 
                          len(stripped_line) > 2 and 
                          not ' ' in stripped_line[2:15]):  # Continuation format
                        
                        # Append continuation line
                        current_dd_statement += " " + stripped_line[2:].strip()
                        
                        # Check if this is the end of the DD statement
                        if not stripped_line.rstrip().endswith(','):
                            self._process_complete_dd_statement(
                                current_dd_statement, file_path, current_dd_line_start,
                                current_job, current_step, current_proc
                            )
                            current_dd_statement = ""
                
                # Process any remaining DD statement
                if current_dd_statement:
                    self._process_complete_dd_statement(
                        current_dd_statement, file_path, current_dd_line_start,
                        current_job, current_step, current_proc
                    )
                        
        except Exception as e:
            print(f"Error analyzing JCL/PROC file {file_path}: {e}")

    def _process_complete_dd_statement(self, dd_statement, file_path, line_number, job_name, step_name, proc_name):
        """Process a complete DD statement (potentially multi-line)."""
        dataset_info = self._parse_dd_statement(
            dd_statement, file_path, line_number, job_name, step_name
        )
        if dataset_info:
            # Add PROC information if available
            if proc_name:
                dataset_info['proc_name'] = proc_name
            else:
                dataset_info['proc_name'] = ''
            
            self.results['jcl_datasets'].append(dataset_info)

    def _parse_dd_statement(self, line, file_path, line_number, job_name, step_name):
        """Parse a DD statement to extract dataset information."""
        try:
            original_line = line.strip()
            stripped_line = original_line.upper()
            
            # Extract DD name - handle both JCL and PROC format
            parts = stripped_line.split()
            if len(parts) < 2:
                return None
            
            # Check if this is a DD statement
            dd_index = -1
            for i, part in enumerate(parts):
                if part == 'DD':
                    dd_index = i
                    break
            
            if dd_index == -1:
                return None
            
            # Extract DD name
            if dd_index == 1 and parts[0].startswith('//'):
                dd_name = parts[0][2:]  # Standard JCL format //DDNAME DD
            elif dd_index == 0:
                dd_name = 'INLINE'  # Just DD parameters
            else:
                dd_name = parts[dd_index - 1]  # Other formats
            
            # Initialize dataset info
            dataset_info = {
                'jcl_file': str(file_path.relative_to(self.root_directory)),
                'job_name': job_name or 'UNKNOWN',
                'step_name': step_name or 'UNKNOWN',
                'dd_name': dd_name,
                'line_number': line_number,
                'dataset_name': '',
                'disp_status': '',
                'disp_normal': '',
                'disp_abnormal': '',
                'dataset_type': '',
                'volume': '',
                'unit': '',
                'space': '',
                'dcb': '',
                'original_line': original_line
            }
            
            # Parse the DD statement parameters - everything after 'DD'
            dd_params = ' '.join(parts[dd_index + 1:])
            
            # Handle parameters that might be quoted
            dd_params = self._clean_jcl_parameters(dd_params)
            
            # Extract DSN (Dataset Name) - handle various formats
            dsn_patterns = [
                r'DSN=([^,\s]+)',           # DSN=DATASET.NAME
                r'DSN=\'([^\']+)\'',        # DSN='DATASET.NAME'
                r'DSN=\"([^\"]+)\"',        # DSN="DATASET.NAME"
                r'DSN=\(([^)]+)\)'          # DSN=(DATASET.NAME)
            ]
            
            for pattern in dsn_patterns:
                dsn_match = re.search(pattern, dd_params)
                if dsn_match:
                    dataset_info['dataset_name'] = dsn_match.group(1).strip('\'"')
                    break
            
            # Extract DISP parameter - handle various formats
            disp_patterns = [
                r'DISP=\(([^)]+)\)',        # DISP=(NEW,CATLG,DELETE)
                r'DISP=([^,\s]+)'           # DISP=SHR
            ]
            
            for pattern in disp_patterns:
                disp_match = re.search(pattern, dd_params)
                if disp_match:
                    disp_value = disp_match.group(1)
                    if ',' in disp_value:
                        disp_parts = [part.strip('\'"') for part in disp_value.split(',')]
                        if len(disp_parts) >= 1:
                            dataset_info['disp_status'] = disp_parts[0].strip()
                        if len(disp_parts) >= 2:
                            dataset_info['disp_normal'] = disp_parts[1].strip()
                        if len(disp_parts) >= 3:
                            dataset_info['disp_abnormal'] = disp_parts[2].strip()
                    else:
                        dataset_info['disp_status'] = disp_value.strip('\'"')
                    break
            
            # Extract other parameters with improved parsing
            param_patterns = {
                'volume': [r'VOL=([^,\s]+)', r'VOLUME=([^,\s]+)', r'VOL=\(([^)]+)\)'],
                'unit': [r'UNIT=([^,\s]+)', r'UNIT=\(([^)]+)\)'],
                'space': [r'SPACE=\(([^)]+(?:\([^)]+\))*[^)]*)\)'],
                'dcb': [r'DCB=\(([^)]+(?:\([^)]+\))*[^)]*)\)']
            }
            
            for param_name, patterns in param_patterns.items():
                for pattern in patterns:
                    match = re.search(pattern, dd_params)
                    if match:
                        dataset_info[param_name] = match.group(1).strip('\'"')
                        break
            
            # Extract SYSOUT parameter
            sysout_match = re.search(r'SYSOUT=([^,\s]+)', dd_params)
            if sysout_match:
                dataset_info['unit'] = f"SYSOUT={sysout_match.group(1)}"
            
            # Determine dataset type with improved logic
            if 'DUMMY' in dd_params:
                dataset_info['dataset_type'] = 'DUMMY'
            elif 'SYSOUT=' in dd_params:
                dataset_info['dataset_type'] = 'SYSOUT'
            elif '*' in dd_params and not dataset_info['dataset_name']:
                dataset_info['dataset_type'] = 'SYSIN'
            elif '&&' in dataset_info.get('dataset_name', ''):
                dataset_info['dataset_type'] = 'TEMPORARY'
            elif dataset_info['dataset_name']:
                dataset_info['dataset_type'] = 'DATASET'
            elif any(keyword in dd_params for keyword in ['PATH=', 'PATHDISP=']):
                dataset_info['dataset_type'] = 'HFS'
            else:
                dataset_info['dataset_type'] = 'OTHER'
            
            # Only return if we found meaningful dataset information
            if (dataset_info['dataset_name'] or 
                dataset_info['disp_status'] or 
                dataset_info['dataset_type'] in ['DUMMY', 'SYSOUT', 'SYSIN']):
                return dataset_info
            
            return None
            
        except Exception as e:
            print(f"Error parsing DD statement at line {line_number}: {e}")
            return None

    def _clean_jcl_parameters(self, params):
        """Clean and normalize JCL parameters for parsing."""
        # Remove extra spaces and normalize
        params = re.sub(r'\s+', ' ', params)
        # Handle continuation characters
        params = params.replace('...', '')
        return params

    def _parse_dd_continuation(self, line, file_path, line_number, job_name, step_name):
        """Parse a DD continuation line to extract dataset information."""
        try:
            stripped_line = line.strip().upper()
            
            # This is a simplified parser for continuation lines
            # In practice, you might need to track the previous DD statement
            dataset_info = {
                'jcl_file': str(file_path.relative_to(self.root_directory)),
                'job_name': job_name or 'UNKNOWN',
                'step_name': step_name or 'UNKNOWN',
                'dd_name': 'CONTINUATION',
                'line_number': line_number,
                'dataset_name': '',
                'disp_status': '',
                'disp_normal': '',
                'disp_abnormal': '',
                'dataset_type': 'CONTINUATION',
                'volume': '',
                'unit': '',
                'space': '',
                'dcb': ''
            }
            
            # Extract DSN from continuation line
            dsn_match = re.search(r'DSN=([^,\s]+)', stripped_line)
            if dsn_match:
                dataset_info['dataset_name'] = dsn_match.group(1)
                dataset_info['dataset_type'] = 'DATASET'
                return dataset_info
            
            # Extract DISP from continuation line
            disp_match = re.search(r'DISP=\(([^)]+)\)', stripped_line)
            if disp_match:
                disp_parts = disp_match.group(1).split(',')
                if len(disp_parts) >= 1:
                    dataset_info['disp_status'] = disp_parts[0].strip()
                if len(disp_parts) >= 2:
                    dataset_info['disp_normal'] = disp_parts[1].strip()
                if len(disp_parts) >= 3:
                    dataset_info['disp_abnormal'] = disp_parts[2].strip()
                return dataset_info
            
            return None
            
        except Exception as e:
            print(f"Error parsing DD continuation at line {line_number}: {e}")
            return None

    def _finalize_analysis(self):
        """Convert sets to lists for JSON serialization and calculate totals."""
        # Convert sets to lists
        self.results['io_analysis']['input_files'] = list(self.results['io_analysis']['input_files'])
        self.results['io_analysis']['output_files'] = list(self.results['io_analysis']['output_files'])
        
        # Calculate detailed line count statistics
        line_stats = {
            'cobol_programs': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            },
            'jcl_files': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            },
            'copybooks': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            },
            'procedures': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            },
            'control_cards': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            },
            'all_files': {
                'total_lines': 0, 'code_lines': 0, 'comment_lines': 0, 'non_empty_lines': 0,
                'files_with_only_comments': 0, 'files_with_code': 0, 'empty_files': 0
            }
        }
        
        # Calculate statistics for each category
        for category in ['cobol_programs', 'jcl_files', 'copybooks', 'procedures', 'control_cards']:
            for file_info in self.results['detailed_files'][category]:
                # Basic line counts
                line_stats[category]['total_lines'] += file_info['lines_of_code']
                line_stats[category]['code_lines'] += file_info['code_lines']
                line_stats[category]['comment_lines'] += file_info['comment_lines']
                line_stats[category]['non_empty_lines'] += file_info['non_empty_lines']
                
                # File classification
                if file_info['lines_of_code'] == 0:
                    line_stats[category]['empty_files'] += 1
                elif file_info['code_lines'] == 0 and file_info['comment_lines'] > 0:
                    line_stats[category]['files_with_only_comments'] += 1
                elif file_info['code_lines'] > 0:
                    line_stats[category]['files_with_code'] += 1
                
                # Add to overall totals
                line_stats['all_files']['total_lines'] += file_info['lines_of_code']
                line_stats['all_files']['code_lines'] += file_info['code_lines']
                line_stats['all_files']['comment_lines'] += file_info['comment_lines']
                line_stats['all_files']['non_empty_lines'] += file_info['non_empty_lines']
            
            # Update overall file classification counts
            line_stats['all_files']['empty_files'] += line_stats[category]['empty_files']
            line_stats['all_files']['files_with_only_comments'] += line_stats[category]['files_with_only_comments']
            line_stats['all_files']['files_with_code'] += line_stats[category]['files_with_code']
        
        # Calculate summary statistics
        self.results['summary'] = {
            'total_cobol_programs': self.results['file_counts']['cobol_programs'],
            'total_jcl_files': self.results['file_counts']['jcl_files'],
            'total_copybooks': self.results['file_counts']['copybooks'],
            'total_procedures': self.results['file_counts']['procedures'],
            'total_control_cards': self.results['file_counts']['control_cards'],
            'total_data_files': self.results['file_counts']['data_files'],
            'total_input_file_references': len(self.results['io_analysis']['input_files']),
            'total_output_file_references': len(self.results['io_analysis']['output_files']),
            'total_jcl_datasets': len(self.results['jcl_datasets']),
            'total_other_files': self.results['file_counts']['other_files'],
            'folders_analyzed': len(self.results['folder_analysis']),
            'line_statistics': line_stats,
            'excluded_files': dict(self.results['excluded_counts']) if self.results['excluded_counts'] else {},
            'total_excluded_files': sum(self.results['excluded_counts'].values()) if self.results['excluded_counts'] else 0
        }

    def generate_report(self, output_format='console'):
        """Generate analysis report in specified format."""
        if output_format == 'console':
            self._print_console_report()
        elif output_format == 'json':
            return self._generate_json_report()
        elif output_format == 'excel':
            return self._generate_excel_report()
        elif output_format == 'csv':
            return self._generate_csv_report()

    def _print_console_report(self):
        """Print formatted report to console."""
        print("\n" + "="*60)
        print("COBOL PROJECT ANALYSIS REPORT")
        print("="*60)
        print(f"Analysis Date: {self.results['scan_timestamp']}")
        print(f"Root Directory: {self.results['root_directory']}")
        print("\n" + "-"*40)
        print("SUMMARY STATISTICS")
        print("-"*40)
        
        summary = self.results['summary']
        print(f"📁 Folders Analyzed: {summary['folders_analyzed']}")
        print(f"💻 COBOL Programs: {summary['total_cobol_programs']}")
        print(f"⚙️  JCL Files: {summary['total_jcl_files']}")
        print(f"📚 Copybooks: {summary['total_copybooks']}")
        print(f"📋 Procedures: {summary['total_procedures']}")
        print(f"🎛️  Control Cards (CTC): {summary['total_control_cards']}")
        print(f"📄 Data Files: {summary['total_data_files']}")
        print(f"📥 Input File References: {summary['total_input_file_references']}")
        print(f"📤 Output File References: {summary['total_output_file_references']}")
        print(f"💾 JCL Datasets: {summary['total_jcl_datasets']}")
        print(f"📎 Other Files: {summary['total_other_files']}")
        
        # Lines of Code Analysis
        if 'line_statistics' in summary:
            print("\n" + "-"*40)
            print("LINES OF CODE ANALYSIS")
            print("-"*40)
            line_stats = summary['line_statistics']
            
            if line_stats['cobol_programs']['total_lines'] > 0:
                stats = line_stats['cobol_programs']
                print(f"💻 COBOL Programs:")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
            
            if line_stats['jcl_files']['total_lines'] > 0:
                stats = line_stats['jcl_files']
                print(f"⚙️  JCL Files:")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
            
            if line_stats['copybooks']['total_lines'] > 0:
                stats = line_stats['copybooks']
                print(f"📚 Copybooks:")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
            
            if line_stats['procedures']['total_lines'] > 0:
                stats = line_stats['procedures']
                print(f"📋 Procedures:")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
            
            if line_stats['control_cards']['total_lines'] > 0:
                stats = line_stats['control_cards']
                print(f"🎛️  Control Cards (CTC):")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
            
            if line_stats['all_files']['total_lines'] > 0:
                stats = line_stats['all_files']
                print(f"📊 Overall Totals:")
                print(f"   Total Lines: {stats['total_lines']:,}")
                print(f"   Code Lines: {stats['code_lines']:,}")
                print(f"   Comment Lines: {stats['comment_lines']:,}")
                print(f"   Non-empty Lines: {stats['non_empty_lines']:,}")
                if stats['files_with_only_comments'] > 0:
                    print(f"   📝 Files with only comments: {stats['files_with_only_comments']}")
                if stats['empty_files'] > 0:
                    print(f"   📄 Empty files: {stats['empty_files']}")
                print(f"   ✅ Files with executable code: {stats['files_with_code']}")
        
        # Excluded files information
        if summary.get('excluded_files') and summary.get('total_excluded_files', 0) > 0:
            print("\n" + "-"*40)
            print("EXCLUDED FILES")
            print("-"*40)
            print(f"🚫 Total Excluded Files: {summary['total_excluded_files']}")
            print("📋 Files excluded because they contain 'THIS JOB IS ALREADY ON LINUX':")
            for file_type, count in summary['excluded_files'].items():
                if count > 0:
                    type_name = self.file_patterns.get(file_type, {}).get('description', file_type.replace('_', ' ').title())
                    print(f"   • {type_name}: {count}")

        # Detailed file breakdown
        self._print_detailed_file_breakdown()
        
        # Folder breakdown
        if self.results['folder_analysis']:
            print("\n" + "-"*40)
            print("FOLDER BREAKDOWN")
            print("-"*40)
            for folder, counts in self.results['folder_analysis'].items():
                if folder == '.':
                    folder = 'Root Directory'
                print(f"\n📂 {folder}:")
                for file_type, count in counts.items():
                    if count > 0:
                        type_name = self.file_patterns.get(file_type, {}).get('description', file_type.replace('_', ' ').title())
                        print(f"   {type_name}: {count}")
        
        # I/O Analysis
        if self.results['io_analysis']['input_files'] or self.results['io_analysis']['output_files']:
            print("\n" + "-"*40)
            print("INPUT/OUTPUT ANALYSIS")
            print("-"*40)
            
            if self.results['io_analysis']['input_files']:
                print(f"\n📥 Input Files Referenced ({len(self.results['io_analysis']['input_files'])}):")
                for file_ref in sorted(self.results['io_analysis']['input_files']):
                    print(f"   • {file_ref}")
            
            if self.results['io_analysis']['output_files']:
                print(f"\n📤 Output Files Referenced ({len(self.results['io_analysis']['output_files'])}):")
                for file_ref in sorted(self.results['io_analysis']['output_files']):
                    print(f"   • {file_ref}")
        
        # JCL Datasets Analysis
        if self.results['jcl_datasets']:
            print("\n" + "-"*40)
            print("JCL DATASETS ANALYSIS")
            print("-"*40)
            
            total_datasets = len(self.results['jcl_datasets'])
            print(f"\n📊 Total Datasets Found: {total_datasets}")
            
            # Count by dataset type
            dataset_types = {}
            disp_statuses = {}
            
            for dataset in self.results['jcl_datasets']:
                dtype = dataset.get('dataset_type', 'UNKNOWN')
                dataset_types[dtype] = dataset_types.get(dtype, 0) + 1
                
                disp = dataset.get('disp_status', 'UNKNOWN')
                disp_statuses[disp] = disp_statuses.get(disp, 0) + 1
            
            print(f"\n📁 Dataset Types:")
            for dtype, count in sorted(dataset_types.items()):
                print(f"   • {dtype}: {count}")
            
            print(f"\n🔧 DISP Status Distribution:")
            for disp, count in sorted(disp_statuses.items()):
                print(f"   • {disp}: {count}")
            
            # Show top datasets by name (if any have names)
            named_datasets = [ds for ds in self.results['jcl_datasets'] if ds.get('dataset_name')]
            if named_datasets:
                print(f"\n📄 Sample Dataset Names:")
                for dataset in named_datasets[:10]:  # Show first 10
                    print(f"   • {dataset['dataset_name']} ({dataset.get('disp_status', 'N/A')})")

    def _print_detailed_file_breakdown(self):
        """Print detailed breakdown of files with line counts."""
        categories_with_files = ['cobol_programs', 'jcl_files', 'copybooks', 'procedures', 'control_cards']
        
        for category in categories_with_files:
            files = self.results['detailed_files'][category]
            if files:
                print(f"\n" + "-"*60)
                category_name = self.file_patterns[category]['description'].upper()
                print(f"DETAILED {category_name} BREAKDOWN")
                print("-"*60)
                
                # Sort files by lines of code (descending)
                sorted_files = sorted(files, key=lambda x: x['lines_of_code'], reverse=True)
                
                for file_info in sorted_files:
                    print(f"📄 {file_info['name']}")
                    print(f"   Path: {file_info['path']}")
                    print(f"   Size: {file_info['size']:,} bytes")
                    print(f"   Total Lines: {file_info['lines_of_code']:,}")
                    print(f"   Code Lines: {file_info['code_lines']:,}")
                    print(f"   Comment Lines: {file_info['comment_lines']:,}")
                    print(f"   Non-empty Lines: {file_info['non_empty_lines']:,}")
                    print()

    def _generate_json_report(self):
        """Generate JSON format report."""
        return json.dumps(self.results, indent=2, default=str)

    def _generate_excel_report(self):
        """Generate Excel format report with multiple sheets."""
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl library is required for Excel export.")
            print("Install it with: pip install openpyxl")
            return None
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"cobol_analysis_{timestamp}.xlsx"
        filepath = self.root_directory / filename
        
        # Create workbook and sheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_detailed_files_sheet(wb)
        self._create_folder_analysis_sheet(wb)
        self._create_io_analysis_sheet(wb)
        
        # Add JCL datasets sheet if we have JCL data
        if self.results['jcl_datasets']:
            self._create_jcl_datasets_sheet(wb)
        
        # Save workbook
        wb.save(filepath)
        print(f"Excel report saved to: {filepath}")
        return filepath

    def _create_summary_sheet(self, workbook):
        """Create summary statistics sheet."""
        ws = workbook.create_sheet("Summary")
        
        # Headers
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Title
        ws['A1'] = "COBOL PROJECT ANALYSIS SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Analysis info
        ws['A3'] = "Analysis Date:"
        ws['B3'] = self.results['scan_timestamp']
        ws['A4'] = "Root Directory:"
        ws['B4'] = self.results['root_directory']
        
        # File counts
        ws['A6'] = "FILE STATISTICS"
        ws['A6'].font = header_font
        
        summary = self.results['summary']
        stats = [
            ("COBOL Programs", summary['total_cobol_programs']),
            ("JCL Files", summary['total_jcl_files']),
            ("Copybooks", summary['total_copybooks']),
            ("Procedures", summary['total_procedures']),
            ("Control Cards (CTC)", summary['total_control_cards']),
            ("Data Files", summary['total_data_files']),
            ("Other Files", summary['total_other_files']),
            ("Folders Analyzed", summary['folders_analyzed']),
            ("Input File References", summary['total_input_file_references']),
            ("Output File References", summary['total_output_file_references'])
        ]
        
        for i, (label, value) in enumerate(stats, start=7):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        # Line statistics
        if 'line_statistics' in summary:
            ws['A16'] = "LINES OF CODE STATISTICS"
            ws['A16'].font = header_font
            
            line_stats = summary['line_statistics']
            
            # Headers for line stats table
            ws['A18'] = "Category"
            ws['B18'] = "Total Lines"
            ws['C18'] = "Code Lines"
            ws['D18'] = "Comment Lines"
            
            for col in ['A18', 'B18', 'C18', 'D18']:
                ws[col].font = header_font
                ws[col].fill = header_fill
            
            categories = [
                ("COBOL Programs", line_stats['cobol_programs']),
                ("JCL Files", line_stats['jcl_files']),
                ("Copybooks", line_stats['copybooks']),
                ("Procedures", line_stats['procedures']),
                ("Control Cards (CTC)", line_stats['control_cards']),
                ("OVERALL TOTAL", line_stats['all_files'])
            ]
            
            for i, (category, stats) in enumerate(categories, start=19):
                ws[f'A{i}'] = category
                ws[f'B{i}'] = stats['total_lines']
                ws[f'C{i}'] = stats['code_lines']
                ws[f'D{i}'] = stats['comment_lines']
                
                if category == "OVERALL TOTAL":
                    for col in ['A', 'B', 'C', 'D']:
                        ws[f'{col}{i}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_detailed_files_sheet(self, workbook):
        """Create detailed files analysis sheet."""
        ws = workbook.create_sheet("Detailed Files")
        
        # Headers
        headers = ["Category", "File Name", "Path", "Size (bytes)", "Total Lines", 
                  "Code Lines", "Comment Lines", "Non-empty Lines"]
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row = 2
        categories = ['cobol_programs', 'jcl_files', 'copybooks', 'procedures', 'control_cards', 'data_files']
        
        for category in categories:
            category_name = self.file_patterns.get(category, {}).get('description', category.replace('_', ' ').title())
            files = self.results['detailed_files'][category]
            
            # Sort by lines of code (descending)
            if files and 'lines_of_code' in files[0]:
                files = sorted(files, key=lambda x: x['lines_of_code'], reverse=True)
            
            for file_info in files:
                ws.cell(row=row, column=1, value=category_name)
                ws.cell(row=row, column=2, value=file_info['name'])
                ws.cell(row=row, column=3, value=file_info['path'])
                ws.cell(row=row, column=4, value=file_info['size'])
                
                if 'lines_of_code' in file_info:
                    ws.cell(row=row, column=5, value=file_info['lines_of_code'])
                    ws.cell(row=row, column=6, value=file_info['code_lines'])
                    ws.cell(row=row, column=7, value=file_info['comment_lines'])
                    ws.cell(row=row, column=8, value=file_info['non_empty_lines'])
                else:
                    for col in range(5, 9):
                        ws.cell(row=row, column=col, value=0)
                
                row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_folder_analysis_sheet(self, workbook):
        """Create folder analysis sheet."""
        ws = workbook.create_sheet("Folder Analysis")
        
        # Headers
        headers = ["Folder", "COBOL Programs", "JCL Files", "Copybooks", "Procedures", "Control Cards", "Data Files", "Other Files"]
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row = 2
        for folder, counts in self.results['folder_analysis'].items():
            folder_name = "Root Directory" if folder == '.' else folder
            ws.cell(row=row, column=1, value=folder_name)
            ws.cell(row=row, column=2, value=counts.get('cobol_programs', 0))
            ws.cell(row=row, column=3, value=counts.get('jcl_files', 0))
            ws.cell(row=row, column=4, value=counts.get('copybooks', 0))
            ws.cell(row=row, column=5, value=counts.get('procedures', 0))
            ws.cell(row=row, column=6, value=counts.get('control_cards', 0))
            ws.cell(row=row, column=7, value=counts.get('data_files', 0))
            ws.cell(row=row, column=8, value=counts.get('other_files', 0))
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_io_analysis_sheet(self, workbook):
        """Create I/O analysis sheet."""
        ws = workbook.create_sheet("IO Analysis")
        
        # Input files section
        ws['A1'] = "INPUT FILES REFERENCED"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for file_ref in sorted(self.results['io_analysis']['input_files']):
            ws.cell(row=row, column=1, value=file_ref)
            row += 1
        
        # Output files section
        output_start_row = row + 2
        ws[f'A{output_start_row}'] = "OUTPUT FILES REFERENCED"
        ws[f'A{output_start_row}'].font = Font(bold=True, size=14)
        
        row = output_start_row + 2
        for file_ref in sorted(self.results['io_analysis']['output_files']):
            ws.cell(row=row, column=1, value=file_ref)
            row += 1
        
        # File references detail
        detail_start_row = row + 2
        ws[f'A{detail_start_row}'] = "FILE REFERENCE DETAILS"
        ws[f'A{detail_start_row}'].font = Font(bold=True, size=14)
        
        # Headers for detail section
        detail_headers = ["File Reference", "Source Program", "Operation Type"]
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        header_row = detail_start_row + 2
        for col, header in enumerate(detail_headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = header_row + 1
        for file_ref, references in self.results['io_analysis']['file_references'].items():
            for ref in references:
                ws.cell(row=row, column=1, value=file_ref)
                ws.cell(row=row, column=2, value=ref['file'])
                ws.cell(row=row, column=3, value=ref['operation'])
                row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_jcl_datasets_sheet(self, workbook):
        """Create JCL datasets analysis sheet."""
        ws = workbook.create_sheet("JCL Datasets")
        
        # Headers
        headers = [
            "JCL File", "Job Name", "Step Name", "PROC Name", "DD Name", "Line Number",
            "Dataset Name", "Dataset Type", "DISP Status", "DISP Normal", 
            "DISP Abnormal", "Volume", "Unit", "Space", "DCB", "Original Line"
        ]
        
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row = 2
        for dataset in self.results['jcl_datasets']:
            ws.cell(row=row, column=1, value=dataset.get('jcl_file', ''))
            ws.cell(row=row, column=2, value=dataset.get('job_name', ''))
            ws.cell(row=row, column=3, value=dataset.get('step_name', ''))
            ws.cell(row=row, column=4, value=dataset.get('proc_name', ''))
            ws.cell(row=row, column=5, value=dataset.get('dd_name', ''))
            ws.cell(row=row, column=6, value=dataset.get('line_number', ''))
            ws.cell(row=row, column=7, value=dataset.get('dataset_name', ''))
            ws.cell(row=row, column=8, value=dataset.get('dataset_type', ''))
            ws.cell(row=row, column=9, value=dataset.get('disp_status', ''))
            ws.cell(row=row, column=10, value=dataset.get('disp_normal', ''))
            ws.cell(row=row, column=11, value=dataset.get('disp_abnormal', ''))
            ws.cell(row=row, column=12, value=dataset.get('volume', ''))
            ws.cell(row=row, column=13, value=dataset.get('unit', ''))
            ws.cell(row=row, column=14, value=dataset.get('space', ''))
            ws.cell(row=row, column=15, value=dataset.get('dcb', ''))
            ws.cell(row=row, column=16, value=dataset.get('original_line', ''))
            row += 1
        
        # Auto-adjust column widths
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            
            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _generate_csv_report(self):
        """Generate CSV format reports (multiple files)."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"cobol_analysis_{timestamp}"
        
        files_created = []
        
        # Summary CSV
        summary_file = self.root_directory / f"{base_filename}_summary.csv"
        self._create_summary_csv(summary_file)
        files_created.append(summary_file)
        
        # Detailed files CSV
        detailed_file = self.root_directory / f"{base_filename}_detailed.csv"
        self._create_detailed_csv(detailed_file)
        files_created.append(detailed_file)
        
        # Folder analysis CSV
        folder_file = self.root_directory / f"{base_filename}_folders.csv"
        self._create_folder_csv(folder_file)
        files_created.append(folder_file)
        
        # I/O analysis CSV
        io_file = self.root_directory / f"{base_filename}_io.csv"
        self._create_io_csv(io_file)
        files_created.append(io_file)
        
        # JCL datasets CSV
        if self.results['jcl_datasets']:
            datasets_file = self.root_directory / f"{base_filename}_jcl_datasets.csv"
            self._create_jcl_datasets_csv(datasets_file)
            files_created.append(datasets_file)
        
        print(f"CSV reports saved:")
        for file in files_created:
            print(f"  - {file}")
        
        return files_created

    def _create_summary_csv(self, filepath):
        """Create summary statistics CSV."""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Header info
            writer.writerow(["COBOL Project Analysis Summary"])
            writer.writerow(["Analysis Date", self.results['scan_timestamp']])
            writer.writerow(["Root Directory", self.results['root_directory']])
            writer.writerow([])  # Empty row
            
            # File statistics
            writer.writerow(["File Statistics"])
            summary = self.results['summary']
            writer.writerow(["Metric", "Count"])
            writer.writerow(["COBOL Programs", summary['total_cobol_programs']])
            writer.writerow(["JCL Files", summary['total_jcl_files']])
            writer.writerow(["Copybooks", summary['total_copybooks']])
            writer.writerow(["Procedures", summary['total_procedures']])
            writer.writerow(["Control Cards (CTC)", summary['total_control_cards']])
            writer.writerow(["Data Files", summary['total_data_files']])
            writer.writerow(["Other Files", summary['total_other_files']])
            writer.writerow(["Folders Analyzed", summary['folders_analyzed']])
            writer.writerow(["Input File References", summary['total_input_file_references']])
            writer.writerow(["Output File References", summary['total_output_file_references']])
            
            # Line statistics
            if 'line_statistics' in summary:
                writer.writerow([])  # Empty row
                writer.writerow(["Lines of Code Statistics"])
                writer.writerow(["Category", "Total Lines", "Code Lines", "Comment Lines"])
                
                line_stats = summary['line_statistics']
                writer.writerow(["COBOL Programs", line_stats['cobol_programs']['total_lines'],
                               line_stats['cobol_programs']['code_lines'], line_stats['cobol_programs']['comment_lines']])
                writer.writerow(["JCL Files", line_stats['jcl_files']['total_lines'],
                               line_stats['jcl_files']['code_lines'], line_stats['jcl_files']['comment_lines']])
                writer.writerow(["Copybooks", line_stats['copybooks']['total_lines'],
                               line_stats['copybooks']['code_lines'], line_stats['copybooks']['comment_lines']])
                writer.writerow(["Procedures", line_stats['procedures']['total_lines'],
                               line_stats['procedures']['code_lines'], line_stats['procedures']['comment_lines']])
                writer.writerow(["Control Cards (CTC)", line_stats['control_cards']['total_lines'],
                               line_stats['control_cards']['code_lines'], line_stats['control_cards']['comment_lines']])
                writer.writerow(["OVERALL TOTAL", line_stats['all_files']['total_lines'],
                               line_stats['all_files']['code_lines'], line_stats['all_files']['comment_lines']])

    def _create_detailed_csv(self, filepath):
        """Create detailed files CSV."""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(["Category", "File Name", "Path", "Size (bytes)", "Total Lines", 
                           "Code Lines", "Comment Lines", "Non-empty Lines"])
            
            # Data
            categories = ['cobol_programs', 'jcl_files', 'copybooks', 'procedures', 'control_cards', 'data_files']
            
            for category in categories:
                category_name = self.file_patterns.get(category, {}).get('description', category.replace('_', ' ').title())
                files = self.results['detailed_files'][category]
                
                # Sort by lines of code if available
                if files and 'lines_of_code' in files[0]:
                    files = sorted(files, key=lambda x: x['lines_of_code'], reverse=True)
                
                for file_info in files:
                    row = [
                        category_name,
                        file_info['name'],
                        file_info['path'],
                        file_info['size']
                    ]
                    
                    if 'lines_of_code' in file_info:
                        row.extend([
                            file_info['lines_of_code'],
                            file_info['code_lines'],
                            file_info['comment_lines'],
                            file_info['non_empty_lines']
                        ])
                    else:
                        row.extend([0, 0, 0, 0])
                    
                    writer.writerow(row)

    def _create_folder_csv(self, filepath):
        """Create folder analysis CSV."""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            writer.writerow(["Folder", "COBOL Programs", "JCL Files", "Copybooks", "Procedures", "Control Cards", "Data Files", "Other Files"])
            
            # Data
            for folder, counts in self.results['folder_analysis'].items():
                folder_name = "Root Directory" if folder == '.' else folder
                writer.writerow([
                    folder_name,
                    counts.get('cobol_programs', 0),
                    counts.get('jcl_files', 0),
                    counts.get('copybooks', 0),
                    counts.get('procedures', 0),
                    counts.get('control_cards', 0),
                    counts.get('data_files', 0),
                    counts.get('other_files', 0)
                ])

    def _create_io_csv(self, filepath):
        """Create I/O analysis CSV."""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Input files section
            writer.writerow(["INPUT FILES REFERENCED"])
            for file_ref in sorted(self.results['io_analysis']['input_files']):
                writer.writerow([file_ref])
            
            writer.writerow([])  # Empty row
            
            # Output files section
            writer.writerow(["OUTPUT FILES REFERENCED"])
            for file_ref in sorted(self.results['io_analysis']['output_files']):
                writer.writerow([file_ref])
            
            writer.writerow([])  # Empty row
            
            # File references detail
            writer.writerow(["FILE REFERENCE DETAILS"])
            writer.writerow(["File Reference", "Source Program", "Operation Type"])
            
            for file_ref, references in self.results['io_analysis']['file_references'].items():
                for ref in references:
                    writer.writerow([file_ref, ref['file'], ref['operation']])

    def _create_jcl_datasets_csv(self, filepath):
        """Create JCL datasets CSV with dataset names and DISP parameters."""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Headers
            headers = [
                "JCL File", "Job Name", "Step Name", "PROC Name", "DD Name", "Line Number",
                "Dataset Name", "Dataset Type", "DISP Status", "DISP Normal", 
                "DISP Abnormal", "Volume", "Unit", "Space", "DCB", "Original Line"
            ]
            writer.writerow(headers)
            
            # Data
            for dataset in self.results['jcl_datasets']:
                writer.writerow([
                    dataset.get('jcl_file', ''),
                    dataset.get('job_name', ''),
                    dataset.get('step_name', ''),
                    dataset.get('proc_name', ''),
                    dataset.get('dd_name', ''),
                    dataset.get('line_number', ''),
                    dataset.get('dataset_name', ''),
                    dataset.get('dataset_type', ''),
                    dataset.get('disp_status', ''),
                    dataset.get('disp_normal', ''),
                    dataset.get('disp_abnormal', ''),
                    dataset.get('volume', ''),
                    dataset.get('unit', ''),
                    dataset.get('space', ''),
                    dataset.get('dcb', ''),
                    dataset.get('original_line', '')
                ])

    def save_report(self, filename=None, format='json'):
        """Save report to file."""
        if format == 'excel':
            return self._generate_excel_report()
        elif format == 'csv':
            return self._generate_csv_report()
        else:  # JSON format
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"cobol_analysis_{timestamp}.{format}"
            
            filepath = self.root_directory / filename
            
            with open(filepath, 'w') as f:
                f.write(self._generate_json_report())
            
            print(f"\nReport saved to: {filepath}")
            return filepath


def main():
    """Main function to run the analysis."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Analyze COBOL project structure and extract statistics")
    parser.add_argument("directory", nargs='?', default=".", 
                       help="Directory to analyze (default: current directory)")
    parser.add_argument("--output", "-o", choices=['console', 'json', 'excel', 'csv', 'all'], 
                       default='console', help="Output format")
    parser.add_argument("--save", "-s", action='store_true', 
                       help="Save report to file")
    parser.add_argument("--export", "-e", choices=['json', 'excel', 'csv', 'all'],
                       help="Export format (alternative to --save with format selection)")
    
    args = parser.parse_args()
    
    # Create analyzer and run analysis
    analyzer = CobolFileAnalyzer(args.directory)
    analyzer.scan_directory()
    
    # Generate report
    if args.output in ['console', 'all']:
        analyzer.generate_report('console')
    
    if args.output == 'json':
        print("\nJSON Report:")
        print(analyzer.generate_report('json'))
    elif args.output == 'excel':
        analyzer.generate_report('excel')
    elif args.output == 'csv':
        analyzer.generate_report('csv')
    elif args.output == 'all':
        print("\nJSON Report:")
        print(analyzer.generate_report('json'))
        analyzer.generate_report('excel')
        analyzer.generate_report('csv')
    
    # Handle save/export options
    if args.save:
        analyzer.save_report(format='json')
    
    if args.export:
        if args.export == 'all':
            analyzer.save_report(format='json')
            analyzer.save_report(format='excel')
            analyzer.save_report(format='csv')
        else:
            analyzer.save_report(format=args.export)


if __name__ == "__main__":
    main() 